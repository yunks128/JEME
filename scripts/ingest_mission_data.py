#!/usr/bin/env python3
"""
Ingest raw GRACE/SWOT publication data into dashboard-ready JSON format.

Reads raw CSV/Excel files from mission_data/, identifies team papers,
links citations to team papers, classifies domains and engagement,
computes uncertainty, and writes to public/data/{MISSION}_analyzed.json.

Usage:
    python scripts/ingest_mission_data.py --mission GRACE
    python scripts/ingest_mission_data.py --mission SWOT
    python scripts/ingest_mission_data.py --all
    python scripts/ingest_mission_data.py --all --dry-run
"""

import argparse
import csv
import hashlib
import json
import re
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Import classification and uncertainty functions from sibling scripts
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR))

from classify_papers import classify_domain, classify_engagement
from compute_uncertainty import compute_entry_uncertainty, MODEL_CORE_KEYWORDS

# Add mission-specific core keywords for uncertainty computation
MODEL_CORE_KEYWORDS["GRACE"] = [
    "grace", "gravity recovery", "grace-fo", "terrestrial water storage",
    "mass change", "gravity field", "tws", "grgs", "csr", "gfz",
    "spherical harmonic", "mascon", "time-variable gravity",
]
MODEL_CORE_KEYWORDS["SWOT"] = [
    "swot", "surface water and ocean topography", "ka-band",
    "karin", "wide-swath altimetry", "water surface elevation",
    "nadir altimetry", "swath altimetry",
]

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TEAM_AFFILIATIONS = [
    "jpl", "jet propulsion laboratory",
    "caltech", "california institute of technology",
]

NULL_SENTINEL = "\\N"

# ---------------------------------------------------------------------------
# JPL Author List (parsed from ~200 JPL scientist profile filenames)
# Each tuple is (last_name_lower, first_name_lower, first_initial_lower)
# ---------------------------------------------------------------------------

_JPL_AUTHOR_FILENAMES = [
    "abhishek-chatterjee", "akiko-hayashi", "alex-gardner", "allison-ho",
    "alyn-lambert", "anamika-shreevastava", "andrew-delman", "angelica-rodriguez",
    "anthony-bloom", "anthony-davis", "anup-nambiathody", "arnaud-cerbelaud",
    "audrey-delpech", "b--jack-pan", "babette-tchonang", "bahruz-ahadov",
    "baijun-tian", "benjamin-hamlington", "benjamin-holt", "bin-guan",
    "bjorn-lambrigtsen", "brandi-downs", "brendan-fisher", "brett-buzzanga",
    "brian-drouin", "brian-kahn", "brian-knosp", "carl-percival",
    "carmen-blackwood--boening-", "carol-bruegge", "cedric-david", "chad-greene",
    "charles-markus", "charles-miller", "cheolhee-yoo", "christiana-ade",
    "christine-lee", "claire-villanueva-weeks", "clement-bertin", "damian-walwer",
    "dana-chadwick", "daniel-cheng", "david-diner", "david-halpern",
    "david-schimel", "deacon-nemchick", "dejian-fu", "denis-demchev",
    "derek-posselt", "duane-waliser", "dustin-roten", "elahe-tajfar",
    "elyse-pennington", "eren-bilir", "eric-fielding", "eric-larour",
    "erik-ivins", "erika-podest", "erin-hightower", "erone-ghizoni-dos-santos",
    "evan-fishbein", "ewa-czyz", "federico-rabuffi", "felix-landerer",
    "fernando-german-chouza-keil", "frank-werner", "frank-winiberg",
    "fredrick-irion", "geoffrey-toon", "gilles-peltzer", "glynn-hulley",
    "graeme-stephens", "gregory-halverson", "gregory-osterman",
    "hanii-takahashi", "hannah-nesser", "heath-rhoades",
    "heidar-thor-thrastarson", "hrusikesha-pradhan", "huanting-huang",
    "ian-fenty", "ian-glenn", "ichiro-fukumori", "ileana-callejas",
    "isabella-velicogna", "isabelle-sanders", "jake-reschke",
    "jean-francois-blavier", "jeffrey-wade", "jeongmin-yun", "jessica-zaiss",
    "jinbo-wang", "jinhyeok-yu", "jinkyul-choi", "joao-teixeira",
    "jonathan-jiang", "jong-hoon-jeong", "josh-willis", "joshua-cuzzone",
    "joshua-laughner", "jt-reager", "juan-crespo", "jui-lin-li",
    "julia-shates", "junjie-liu", "kathleen--katie--grant",
    "kazuyuki-miyazaki", "keeyoon-sung", "kelly-luis",
    "kerry-cawse-nicholson", "kevin-bowman", "kevin-schwarm",
    "kyeore--holly--han", "kyle-mcdonald", "kyra-adams--kim-",
    "kyriaki-drymoni", "lambert-caron", "lance-christensen", "le-kuai",
    "leon-maldonado", "lucien-froidevaux", "luis-mill-n", "luke-kachelein",
    "madeleine-pascolini-campbell", "mandy-lopez", "marcin-kurowski",
    "marcin-witek", "maria-chinita", "maria-hakuba", "marie-zahn",
    "mark-richardson", "mark-smalley", "marufa-ishaque", "matth-us-kiel",
    "matthew-archer", "matthew-bonnema", "matthew-lebsock", "mattia-poinelli",
    "michael-abrams", "michael-keller", "michael-schwartz", "michelle-gierach",
    "michelle-santee", "nathaniel-livesey", "naveen-ramachandran",
    "nicholas-parazoo", "nima-madani", "noah-molotch",
    "odilon-houndegnonto", "olga-kalashnikova", "ou-wang",
    "oumaima-lamaakel", "patrick-wang", "paul-levine", "paul-lundgren",
    "qing-yue", "rapha-l-savelli", "rashmi-shah", "renato-braghiere",
    "renato-prata-de-moraes-frasson", "rishav-mallick", "robert-herman",
    "robert-nelson", "robert-radocinski", "ryan-fuller", "sahra-kacimi",
    "sarah-worden", "sassan-saatchi", "severine-fournier", "shakeel-asharaf",
    "shuang-ma", "simon-yueh", "sina-hasheminassab", "sreelekha-jarugula",
    "sudhanshu-pandey", "sun-wong", "surendra-adhikari", "susan-owen",
    "tao-wang", "terry-kubar", "thierry-leblanc", "thomas-kurosu",
    "thomas-pongetti", "tianlin-wang", "timothy-crawford", "tong-lee",
    "vicky-espinoza", "victor-zlotnicki", "vijay-natraj",
    "vincent-realmuto", "vivienne-payne", "wen-chao", "xianan-jiang",
    "xiaochun--adam--wang", "yingdi-luo", "yuhe-song", "zhen-liu",
]


def _build_jpl_author_lookup():
    """Build a lookup dict: last_name_lower -> set of first_initials_lower.

    Handles special filename formats like:
      - "b--jack-pan" -> last=pan, firsts={b, jack, j}
      - "carmen-blackwood--boening-" -> last=boening, firsts={carmen, c}
      - "kathleen--katie--grant" -> last=grant, firsts={kathleen, katie, k}
      - "xiaochun--adam--wang" -> last=wang, firsts={xiaochun, adam, x, a}
    """
    lookup = {}  # last_name -> set of (first_name, first_initial)

    for raw in _JPL_AUTHOR_FILENAMES:
        # Split on single hyphens (not double-hyphens which are separators/parens)
        # First normalize: replace -- with a special marker
        normalized = raw.replace("--", "||")
        # Remove trailing markers (from parenthetical suffixes like "boening-")
        normalized = normalized.strip("|").strip("-")
        parts = [p.strip("-") for p in normalized.split("||") if p.strip("-")]

        if not parts:
            continue

        # The structure is: first-...-last or first||middle||last
        # For simple names like "felix-landerer": parts=["felix-landerer"]
        # For names like "kathleen||katie||grant": parts=["kathleen","katie","grant"]
        # For "b||jack-pan": parts=["b","jack-pan"]
        # For "carmen-blackwood||boening": parts=["carmen-blackwood","boening"]

        # Collect all name segments by splitting remaining hyphens
        all_segments = []
        for part in parts:
            all_segments.extend(part.split("-"))
        all_segments = [s for s in all_segments if s]

        if len(all_segments) < 2:
            continue

        # Last segment is last name
        last = all_segments[-1].lower()
        firsts = set()
        for seg in all_segments[:-1]:
            seg_lower = seg.lower()
            firsts.add(seg_lower)
            firsts.add(seg_lower[0])  # first initial

        if last not in lookup:
            lookup[last] = set()
        lookup[last].update(firsts)

    return lookup


# Pre-build the lookup at module load time
_JPL_AUTHOR_LOOKUP = _build_jpl_author_lookup()


def is_null(val):
    """Check if a raw CSV/Excel value is null."""
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s == NULL_SENTINEL or s == "None"


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_grace_csv(filepath):
    """Parse GRACE tab-delimited CSV into list of dicts."""
    records = []
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            records.append(row)
    print(f"  Parsed {len(records)} rows from GRACE CSV")
    return records


def parse_swot_excel(filepath):
    """Parse SWOT Excel Publications_All_Time sheet into list of dicts."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["Publications_All_Time"]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for h, v in zip(header, row):
            if h is not None:
                record[h] = v
        # Skip rows that are summaries or empty
        if record.get("id") is not None and record.get("title") is not None:
            records.append(record)

    wb.close()
    print(f"  Parsed {len(records)} rows from SWOT Excel")
    return records


# ---------------------------------------------------------------------------
# Team paper identification
# ---------------------------------------------------------------------------

def _is_team_by_address(record):
    """Check if the paper is a team paper based on author_address field."""
    addr = str(record.get("author_address", "") or "").lower()
    for affiliation in TEAM_AFFILIATIONS:
        if affiliation in addr:
            return True
    return False


def _author_matches_jpl(last, first_token, jpl_firsts, strict):
    """Check if a parsed author name matches a JPL entry.

    Args:
        last: lowercase last name
        first_token: lowercase first name/initial token (e.g. "felix", "f")
        jpl_firsts: set of known JPL first names/initials for this last name
        strict: if True, require full first name match (not just initial)
    """
    if not first_token:
        return False

    # Full first name match (e.g. "felix" in {"f", "felix"})
    if len(first_token) > 1 and first_token in jpl_firsts:
        return True

    # Initial-only match only for non-strict surnames
    if not strict:
        first_initial = first_token[0]
        if first_initial in jpl_firsts:
            return True

    return False


# Common surnames where initial-only matching causes too many false positives.
# For these, we require the full first name to match.
_STRICT_MATCH_SURNAMES = {
    "wang", "liu", "li", "lee", "kim", "pan", "ma", "fu", "yu", "choi",
    "yoo", "wu", "chen", "zhang", "yang", "huang", "song", "chao",
    "guan", "yue", "jiang",
}


def _is_team_by_author(record):
    """Check if any author matches the JPL scientist list.

    Handles author formats like:
      - "Landerer, Felix W."  -> last=landerer, match felix
      - "Landerer, F."        -> last=landerer, match initial f
      - "Walwer, D."          -> last=walwer, match initial d
      - "F. Landerer"         -> last=landerer, match initial f
    For common surnames (Wang, Liu, etc.), requires full first name match
    to avoid false positives.
    """
    authors_raw = record.get("authors", "")
    if is_null(authors_raw):
        return False

    authors = [a.strip() for a in str(authors_raw).split(";") if a.strip()]

    for author in authors:
        # Try "Last, First ..." format
        if "," in author:
            parts = author.split(",", 1)
            last = parts[0].strip().lower().strip(".")
            first_part = parts[1].strip().lower() if len(parts) > 1 else ""
        else:
            # Try "First Last" format
            tokens = author.strip().split()
            if len(tokens) < 2:
                continue
            last = tokens[-1].strip(".").lower()
            first_part = tokens[0].strip(".").lower()

        if last not in _JPL_AUTHOR_LOOKUP:
            continue

        # Extract first name token (handles "Felix W." -> "felix")
        if not first_part:
            continue
        first_token = first_part.split()[0].strip(".").lower()
        if not first_token:
            continue

        strict = last in _STRICT_MATCH_SURNAMES
        jpl_firsts = _JPL_AUTHOR_LOOKUP[last]

        if _author_matches_jpl(last, first_token, jpl_firsts, strict):
            return True

    return False


def is_team_paper(record):
    """Check if the paper is a team (JPL/Caltech) paper.

    Uses two signals:
    1. author_address contains JPL/Caltech affiliation
    2. Any author name matches the JPL scientist list (~200 names)
    """
    return _is_team_by_address(record) or _is_team_by_author(record)


# ---------------------------------------------------------------------------
# Venue extraction
# ---------------------------------------------------------------------------

def extract_venue(record):
    """Extract journal/venue name from the citation field."""
    citation = str(record.get("citation", "") or "")
    if not citation:
        return ""

    # Typical citation format:
    # Author(s) (Year), Title, Journal, Volume(Issue), Pages, doi
    # or: Author(s), Title, Journal, Volume, Pages
    # Try to extract journal name after the title portion

    # Strategy: split on commas, find the part after year that looks like a journal
    # Remove the leading author+year portion
    # Look for pattern: ), Journal Name,
    match = re.search(r'\)\s*,\s*(.+)', citation)
    if match:
        remainder = match.group(1)
        # Split remainder by comma, journal is typically the 2nd part (after title)
        parts = [p.strip() for p in remainder.split(",")]
        if len(parts) >= 2:
            # First part is title, second is journal
            candidate = parts[1].strip()
            # Clean up volume/issue/pages
            candidate = re.sub(r'\s*\d+\(.*$', '', candidate)
            candidate = re.sub(r'\s*vol\.?\s*\d+.*$', '', candidate, flags=re.IGNORECASE)
            if candidate and len(candidate) > 3:
                return candidate

    return ""


# ---------------------------------------------------------------------------
# Field mapping
# ---------------------------------------------------------------------------

def make_paper_id(record, mission_name):
    """Generate a stable paper_id from the record."""
    ext_uid = record.get("external_uid")
    if not is_null(ext_uid):
        return f"{mission_name.lower()}_{ext_uid}"

    # Fallback: hash of title
    title = str(record.get("title", ""))
    return hashlib.sha256(title.encode("utf-8")).hexdigest()[:40]


def parse_authors(raw):
    """Split author string (semicolon-delimited) into a list."""
    if is_null(raw):
        return []
    raw = str(raw)
    # Handle both "; " and ";" separators
    authors = [a.strip() for a in re.split(r";\s*", raw) if a.strip()]
    return authors


def parse_year(raw):
    """Parse publication year."""
    if is_null(raw):
        return None
    try:
        return int(float(str(raw)))
    except (ValueError, TypeError):
        return None


def map_to_dashboard_schema(record, mission_name, is_team, team_paper_title=None, team_paper_id=None):
    """Convert a raw record to the dashboard JSON schema."""
    title = str(record.get("title", "")).strip()
    doi_raw = record.get("doi")
    doi = "" if is_null(doi_raw) else str(doi_raw).strip()
    link = str(record.get("link", "") or "").strip()
    abstract = None if is_null(record.get("abstract")) else str(record["abstract"]).strip()

    entry = {
        "title": title,
        "authors": parse_authors(record.get("authors")),
        "year": parse_year(record.get("publication_year")),
        "doi": doi,
        "venue": extract_venue(record),
        "citation_count": 0,
        "paper_id": make_paper_id(record, mission_name),
        "url": link if link and link != NULL_SENTINEL else "",
        "abstract": abstract,
        "citing_team_paper": team_paper_title if not is_team else None,
        "team_paper_id": team_paper_id if not is_team else None,
        "research_domain": None,  # filled later
        "engagement_level": None,  # filled later
        "missions_instruments": [
            {
                "name": mission_name,
                "agency": "NASA",
                "type": "Satellite",
                "product": "Not specified",
                "data_level": "Not specified",
                "usage_context": f"{'Team publication for' if is_team else 'Cites'} {mission_name} mission",
            }
        ],
    }

    # For GRACE, add GRACE-FO if mentioned in title/abstract
    text = f"{title} {abstract or ''}".lower()
    if mission_name == "GRACE" and ("grace-fo" in text or "grace follow" in text):
        entry["missions_instruments"].append({
            "name": "GRACE-FO",
            "agency": "NASA",
            "type": "Satellite",
            "product": "Not specified",
            "data_level": "Not specified",
            "usage_context": "Related GRACE Follow-On mission",
        })

    return entry


# ---------------------------------------------------------------------------
# Team paper linking
# ---------------------------------------------------------------------------

def build_keyword_index(team_papers):
    """Build a keyword set for each team paper from title + abstract."""
    stop_words = {
        "the", "a", "an", "and", "or", "of", "in", "to", "for", "with",
        "on", "at", "by", "from", "is", "are", "was", "were", "be", "been",
        "its", "their", "this", "that", "these", "those", "using", "based",
    }
    index = []
    for tp in team_papers:
        text = f"{tp.get('title', '')} {tp.get('abstract', '') or ''}".lower()
        words = set(re.findall(r"\b[a-z]{3,}\b", text)) - stop_words
        index.append((tp, words))
    return index


def find_best_team_paper(citation_record, team_index):
    """Find the best-matching team paper for a citation using keyword overlap."""
    if not team_index:
        return None, None

    text = f"{citation_record.get('title', '')} {citation_record.get('abstract', '') or ''}".lower()
    stop_words = {
        "the", "a", "an", "and", "or", "of", "in", "to", "for", "with",
        "on", "at", "by", "from", "is", "are", "was", "were", "be", "been",
        "its", "their", "this", "that", "these", "those", "using", "based",
    }
    cite_words = set(re.findall(r"\b[a-z]{3,}\b", text)) - stop_words

    best_score = 0
    best_tp = None
    for tp, tp_words in team_index:
        if not tp_words or not cite_words:
            continue
        overlap = len(cite_words & tp_words)
        # Jaccard-like score normalized by smaller set
        score = overlap / min(len(cite_words), len(tp_words)) if min(len(cite_words), len(tp_words)) > 0 else 0
        if score > best_score:
            best_score = score
            best_tp = tp

    if best_tp and best_score > 0.1:
        return str(best_tp.get("title", "")), make_paper_id(best_tp, "")
    return None, None


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def process_mission(mission_name, project_root, dry_run=False):
    """Process a single mission's raw data into dashboard JSON."""
    mission_data_dir = project_root / "mission_data"
    output_dir = project_root / "public" / "data"

    print(f"\n[{mission_name}]")

    # 1. Parse raw data
    if mission_name == "GRACE":
        raw_records = parse_grace_csv(mission_data_dir / "grace.publications.csv")
    elif mission_name == "SWOT":
        raw_records = parse_swot_excel(mission_data_dir / "SWOT_Publications_20250912.xlsx")
    else:
        print(f"  ERROR: Unknown mission '{mission_name}'")
        return None

    # 2. Split into team papers vs citation papers
    team_raw = [r for r in raw_records if is_team_paper(r)]
    cite_raw = [r for r in raw_records if not is_team_paper(r)]
    print(f"  Team papers: {len(team_raw)}, Citation papers: {len(cite_raw)}")

    # 3. Build keyword index for team papers
    team_index = build_keyword_index(team_raw)

    # 4. Map team papers
    entries = []
    for r in team_raw:
        entry = map_to_dashboard_schema(r, mission_name, is_team=True)
        entries.append(entry)

    # 5. Map citation papers with team paper linking
    linked_count = 0
    for r in cite_raw:
        tp_title, tp_id = find_best_team_paper(r, team_index)
        entry = map_to_dashboard_schema(
            r, mission_name, is_team=False,
            team_paper_title=tp_title, team_paper_id=tp_id,
        )
        if tp_title:
            linked_count += 1
        entries.append(entry)

    print(f"  Linked {linked_count}/{len(cite_raw)} citations to team papers")

    # 6. Filter invalid years
    MIN_YEAR = 1990
    MAX_YEAR = 2026
    original_count = len(entries)
    entries = [e for e in entries if e["year"] is None or (MIN_YEAR <= e["year"] <= MAX_YEAR)]
    removed = original_count - len(entries)
    if removed:
        print(f"  Removed {removed} entries with invalid years")

    # 7. Classify domain and engagement
    domain_counts = {}
    engagement_counts = {}
    for entry in entries:
        domain = classify_domain(entry)
        entry["research_domain"] = domain
        domain_counts[domain] = domain_counts.get(domain, 0) + 1

        engagement = classify_engagement(entry)
        entry["engagement_level"] = engagement
        engagement_counts[engagement] = engagement_counts.get(engagement, 0) + 1

    print(f"  Domain distribution: {domain_counts}")
    print(f"  Engagement distribution: {engagement_counts}")

    # 8. Compute uncertainty
    for entry in entries:
        entry["uncertainty"] = compute_entry_uncertainty(entry, mission_name)

    # Compute stats
    high = sum(1 for e in entries if e["uncertainty"]["composite_confidence"] >= 0.7)
    med = sum(1 for e in entries if 0.4 <= e["uncertainty"]["composite_confidence"] < 0.7)
    low = sum(1 for e in entries if e["uncertainty"]["composite_confidence"] < 0.4)
    with_abstract = sum(1 for e in entries if e.get("abstract"))
    print(f"  Uncertainty: high={high}, medium={med}, low={low}")
    print(f"  Abstract coverage: {with_abstract}/{len(entries)} ({100*with_abstract/len(entries):.1f}%)")

    # 9. Write output
    output_path = output_dir / f"{mission_name}_analyzed.json"
    if dry_run:
        print(f"  DRY RUN: would write {len(entries)} entries to {output_path}")
        if entries:
            print(f"  Sample entry:")
            print(json.dumps(entries[0], indent=4)[:1000])
    else:
        output_dir.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(entries, f, indent=2, ensure_ascii=False)
        print(f"  Written {len(entries)} entries to {output_path}")

    return {
        "total": len(entries),
        "team_papers": len(team_raw),
        "citation_papers": len(cite_raw),
        "linked": linked_count,
        "with_abstract": with_abstract,
    }


def main():
    parser = argparse.ArgumentParser(description="Ingest GRACE/SWOT mission data")
    parser.add_argument("--mission", type=str, help="Process a specific mission (GRACE or SWOT)")
    parser.add_argument("--all", action="store_true", help="Process all missions")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing files")
    args = parser.parse_args()

    if not args.mission and not args.all:
        parser.print_help()
        sys.exit(1)

    project_root = SCRIPT_DIR.parent

    missions = ["GRACE", "SWOT"]
    if args.mission:
        if args.mission not in missions:
            print(f"ERROR: Unknown mission '{args.mission}'. Valid: {', '.join(missions)}")
            sys.exit(1)
        missions = [args.mission]

    print(f"Ingesting mission data {'(DRY RUN)' if args.dry_run else ''}")

    for mission in missions:
        stats = process_mission(mission, project_root, dry_run=args.dry_run)
        if stats:
            print(f"  Summary: {stats}")

    print("\nDone.")


if __name__ == "__main__":
    main()
